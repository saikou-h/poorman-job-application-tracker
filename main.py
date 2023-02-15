# -*- coding: utf-8 -*-
"""
Created on Mon Feb 13 12:57:45 2023

Poor man's job application tracker

@author: Higo
"""

import sys, os
from pathlib import Path
from PyQt5 import QtWidgets, uic
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import openpyxl
import pandas as pd
from dataframe_model import PandasModel


class Ui_MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(Ui_MainWindow, self).__init__()
        self.path = Path(__file__).parent.resolve()
        self.log_filename = 'Tracker.xlsx'
        uic.loadUi(self.path / 'main.ui', self)
        
        # setting up
        self.setWindowTitle("Job application tracker")
        self.submissionDate.setDate(QDate.currentDate())
        self.tv = QtWidgets.QTableView()
        
        # calling method
        self.ui_components()
        
        self.show()
        
    def ui_components(self):
        # add_button = QPushButton("addButton", self)
        self.addButton.clicked.connect(self.add_to_log)
        self.resetButton.clicked.connect(self.reset_ui)
        self.reviewButton.clicked.connect(self.review_application)
        
    
    def add_to_log(self):
        print("Add button clicked.")
        text_company = self.companyText.text()
        text_position = self.positionText.text()
        text_link = self.linkText.text()
        text_worktype = self.worktypeText.text()
        text_location = self.locationText.text()
        text_salary = self.salaryText.text()
        text_submitted = str(self.submissionCombo.currentText())
        text_submission = self.submissionDate.text()
        text_source = self.sourceText.text()
        text_note = self.noteText.text()
        print("Job info: \n    \
              Company: {}\n    \
              Position: {}\n   \
              Job Link: {}\n    \
              Work Type: {}\n    \
              Location: {}\n    \
              Salary: {}\n    \
              Submitted: {}\n    \
              Submission Date: {}\n    \
              Source: {}\n    \
              Note: {}".format(
                  text_company, text_position, text_link, text_worktype,
                  text_location, text_salary, text_submitted, text_submission,
                  text_source, text_note
              ))
    
        log_file = openpyxl.load_workbook(self.path / self.log_filename)
        log_sheet = log_file['Sheet1']
        curr_index = len(log_sheet['A'])
        print(curr_index)
        log_sheet.cell(row=curr_index+1, column=1).value = curr_index
        log_sheet.cell(row=curr_index+1, column=2).value = text_company
        log_sheet.cell(row=curr_index+1, column=3).value = text_position
        log_sheet.cell(row=curr_index+1, column=3).hyperlink = text_link
        log_sheet.cell(row=curr_index+1, column=3).style = "Hyperlink"
        log_sheet.cell(row=curr_index+1, column=4).value = text_worktype
        log_sheet.cell(row=curr_index+1, column=5).value = text_location
        log_sheet.cell(row=curr_index+1, column=6).value = text_salary
        log_sheet.cell(row=curr_index+1, column=7).value = text_submitted
        log_sheet.cell(row=curr_index+1, column=8).value = text_submission
        log_sheet.cell(row=curr_index+1, column=8).number_format = 'm/d/yyyy'
        log_sheet.cell(row=curr_index+1, column=8).alignment \
            = openpyxl.styles.Alignment(horizontal='right')
        log_sheet.cell(row=curr_index+1, column=9).value = text_source
        log_sheet.cell(row=curr_index+1, column=10).value = text_note
        log_file.save(self.path / self.log_filename)
        QtWidgets.QMessageBox.about(
            self, "Confirmation", "Application to {} at {} has been saved!".format(
                text_position, text_company    
            )
        )
        
        
    def reset_ui(self):
        print("Reset button clicked.")
        self.companyText.setText("")
        self.positionText.setText("")
        self.linkText.setText("")
        self.worktypeText.setText("")
        self.locationText.setText("")
        self.salaryText.setText("")
        self.submissionCombo.setCurrentIndex(0)
        self.submissionDate.setDate(QDate.currentDate())
        self.sourceText.setText("")      
        self.noteText.setText("")
    
    def review_application(self):
        """
        To be implemented. This button displays a dataframe with entries 
        currently in the tracker spreadsheet

        Returns
        -------
        None.

        """
        
        

if __name__ == "__main__":
    new_app = QtWidgets.QApplication(sys.argv)
    new_window = Ui_MainWindow()
    new_app.exec_()
    
